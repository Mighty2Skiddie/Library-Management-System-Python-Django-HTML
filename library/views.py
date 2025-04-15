from django.views.generic import CreateView, ListView,View
from django.urls import reverse_lazy
from django.shortcuts import render
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.http import HttpResponse
from openpyxl import Workbook

def home(request):
    return render(request, 'library/home.html')

# Create new Author
class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'library/author_form.html'
    success_url = reverse_lazy('author-list')

# Create new Book
class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library/book_form.html'
    success_url = reverse_lazy('book-list')

# Create new Borrow Record
class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'library/borrow_form.html'
    success_url = reverse_lazy('borrow-list')
class AuthorListView(ListView):
    model = Author
    template_name = 'library/author_list.html'
    context_object_name = 'authors'
    paginate_by = 5

class BookListView(ListView):
    model = Book
    template_name = 'library/book_list.html'
    context_object_name = 'books'
    paginate_by = 5

class BorrowRecordListView(ListView):
    model = BorrowRecord
    template_name = 'library/borrow_list.html'
    context_object_name = 'borrow_records'
    paginate_by = 5

class ExportToExcelView(View):
    def get(self, request):
        wb = Workbook()

        # Authors Sheet
        ws1 = wb.active
        ws1.title = "Authors"
        ws1.append(['ID', 'Name', 'Email', 'Bio'])
        for author in Author.objects.all():
            ws1.append([author.id, author.name, author.email, author.bio])

        # Books Sheet
        ws2 = wb.create_sheet(title="Books")
        ws2.append(['ID', 'Title', 'Genre', 'Published Date', 'Author'])
        for book in Book.objects.all():
            ws2.append([book.id, book.title, book.genre, book.published_date, book.author.name])

        # Borrow Records Sheet
        ws3 = wb.create_sheet(title="BorrowRecords")
        ws3.append(['ID', 'User Name', 'Book', 'Borrow Date', 'Return Date'])
        for record in BorrowRecord.objects.all():
            ws3.append([record.id, record.user_name, record.book.title, record.borrow_date, record.return_date])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
        wb.save(response)
        return response
    
    